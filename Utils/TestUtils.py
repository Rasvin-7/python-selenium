import openpyxl
from openpyxl import load_workbook


class TestUtils:
    login = "Login"

    workbook = load_workbook("C:/Users/developer/PycharmProjects/SeleniumPytest/Utils/OrangeHRM TestData.xlsx")

    def getMaxRow(self):
        sheet = self.workbook[self.login]
        return sheet.max_row

    def getMaxCol(self):
        sheet = self.workbook[self.login]
        return sheet.max_column

    def getData(self):
        sheet = self.workbook[self.login]
        cell = []
        cell1 = ()
        data1 = []
        rows = self.getMaxRow()
        cols = self.getMaxCol()
        # for row in sheet.iter_rows():
        #     for cellv in row:
        #         cell.append(cellv.value)
        #     data1.append(tuple(cell))
        #     cell.clear()
        #     print(data1)

        for row in range(2, rows + 1):
            for col in range(1, cols + 1):
                data = sheet.cell(row, col).value
                print(data, "\n")
                cell.append(data)

            cell1 = tuple(cell)
            cell.clear()
            print(cell1, "\n")
            data1.append(cell1)
        return data1


tu = TestUtils()
val = tu.getData()
print(val)
